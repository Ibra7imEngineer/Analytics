"""
Universal Data Extraction Portal - Professional Edition
=======================================================
Commercial-grade Streamlit backend for a dynamic page fetcher, AI-driven schema extraction,
and export-ready data pipeline with Playwright and Google Gemini integration.
"""

import io
import json
import random
import re
from typing import Any, Dict, List, Tuple

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup

try:
    import google.generativeai as genai
    GENAI_AVAILABLE = True
except ImportError:
    GENAI_AVAILABLE = False

try:
    from playwright.sync_api import Error as PlaywrightError
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = Font = PatternFill = Alignment = None

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEFAULT_TIMEOUT = 55
MAX_FILE_SIZE = 20 * 1024 * 1024
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.6420.93 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.5 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 18_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.0 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0",
]
GEMINI_API_BASE = "https://generativelanguage.googleapis.com/v1beta/models"
DEFAULT_GEMINI_MODEL = "gemini-1.5-flash"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_text(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text.strip())
    return ''.join(ch for ch in text if ch.isprintable())


def normalize_api_key(raw_key: str) -> str:
    if not raw_key:
        return ""
    raw = raw_key.strip()
    raw = re.sub(r"[\s]+", " ", raw)
    api_key_match = re.search(r"(ya29\.[A-Za-z0-9\-_]+|AIza[0-9A-Za-z\-_]{35,45})", raw)
    if api_key_match:
        return api_key_match.group(1)
    return raw


def is_valid_api_key(api_key: str) -> bool:
    if not api_key:
        return False
    if re.search(r"[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]", api_key):
        return False
    if " " in api_key or "\n" in api_key or "\r" in api_key:
        return False
    return bool(re.match(r"^(ya29\.[A-Za-z0-9\-_]+|AIza[0-9A-Za-z\-_]{35,45})$", api_key))


def validate_url(url: str) -> bool:
    try:
        result = requests.utils.urlparse(url)
        return all([result.scheme in ("http", "https"), result.netloc])
    except Exception:
        return False


def is_blocked_page(text: str) -> bool:
    if not text:
        return True
    lower = text.lower()
    patterns = [
        "access denied",
        "forbidden",
        "captcha",
        "bot detection",
        "are you human",
        "cloudflare",
        "request blocked",
        "browser verification",
        "verify you are",
        "security check",
    ]
    return any(pattern in lower for pattern in patterns) or len(text) < 40


def remove_noise(soup: BeautifulSoup) -> None:
    for tag in soup(["script", "style", "noscript", "iframe", "header", "footer", "nav", "aside", "form", "svg", "button"]):
        tag.decompose()


def simplify_html_to_text(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    remove_noise(soup)
    raw_text = soup.get_text(separator=" ", strip=True)
    return clean_text(raw_text)


def extract_json_array(raw: str) -> List[Dict[str, Any]]:
    raw = raw.strip()
    if not raw:
        return []

    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        start = raw.find("[")
        end = raw.rfind("]")
        if start >= 0 and end > start:
            try:
                parsed = json.loads(raw[start:end + 1])
            except json.JSONDecodeError:
                parsed = None
        else:
            parsed = None

    if parsed is None:
        return []

    if isinstance(parsed, dict):
        for candidate in ("items", "data", "records", "rows", "results"):
            if candidate in parsed and isinstance(parsed[candidate], list):
                return parsed[candidate]
        return [parsed]

    if isinstance(parsed, list):
        return [item for item in parsed if isinstance(item, dict)]

    return []


def build_gemini_prompt(page_text: str) -> str:
    cleaned_page_text = page_text[:150000]  # Increased limit for more context
    return (
        "You are an expert AI data extraction specialist. Your task is to analyze the provided webpage content and extract structured data into a JSON array of objects.\n\n"
        "INSTRUCTIONS:\n"
        "- Identify the main repeating entities (products, articles, listings, records, items, etc.)\n"
        "- Create descriptive field names based on the content (e.g., 'name', 'price', 'description', 'date', etc.)\n"
        "- Extract ALL available fields for each entity - be comprehensive\n"
        "- If data appears in tables, lists, or repeated patterns, extract it\n"
        "- Handle various formats: prices with currency, dates, numbers, text\n"
        "- Return ONLY a valid JSON array of objects, no explanations or markdown\n"
        "- If no structured data found, return empty array: []\n"
        "- Ensure field names are in English but preserve original language content\n"
        "- Be thorough: extract contact info, addresses, specifications, reviews, etc. if present\n\n"
        "EXAMPLE OUTPUT FORMAT:\n"
        '[{"name": "Product A", "price": "$10.99", "category": "Electronics"}, {"name": "Product B", "price": "$15.50", "category": "Books"}]\n\n'
        "WEBPAGE CONTENT:\n" + cleaned_page_text
    )


def parse_gemini_response(response_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    candidates = response_json.get("candidates") or response_json.get("outputs") or []
    if isinstance(candidates, list) and candidates:
        candidate = candidates[0]
        text = candidate.get("output") or candidate.get("content") or ""
        if isinstance(text, list):
            text = " ".join(str(item.get("text", "")) if isinstance(item, dict) else str(item) for item in text)
        if isinstance(text, dict):
            text = text.get("text", "")
        if isinstance(text, str) and text.strip():
            return extract_json_array(text)

    if "output" in response_json and isinstance(response_json["output"], str):
        return extract_json_array(response_json["output"])

    return []


def parse_gemini_response_from_genai(response) -> List[Dict[str, Any]]:
    text = response.text
    return extract_json_array(text)


def list_gemini_models(api_key: str, timeout: int = 30) -> List[str]:
    if not GENAI_AVAILABLE:
        return []
    try:
        genai.configure(api_key=api_key)
        models = genai.list_models()
        return [model.name for model in models if 'generateContent' in model.supported_generation_methods]
    except Exception:
        return []


def call_gemini_api(page_text: str, api_key: str, model: str = DEFAULT_GEMINI_MODEL, timeout: int = 55) -> Tuple[List[Dict[str, Any]], str]:
    if not GENAI_AVAILABLE:
        raise ImportError("google-generativeai library is not installed.")
    if not api_key:
        raise ValueError("Gemini API key is required.")

    genai.configure(api_key=api_key)
    candidate_models = [model, "gemini-1.5-flash", "gemini-1.5-pro", "gemini-1.0-pro"]
    candidate_models = list(dict.fromkeys([m for m in candidate_models if m]))

    prompt = build_gemini_prompt(page_text)

    last_error = None
    for candidate_model in candidate_models:
        try:
            model_instance = genai.GenerativeModel(candidate_model)
            response = model_instance.generate_content(prompt, generation_config=genai.types.GenerationConfig(
                temperature=0.0,
                max_output_tokens=1200,
                top_p=0.95,
                top_k=40,
            ))
            return parse_gemini_response_from_genai(response), candidate_model
        except Exception as exc:
            last_error = exc
            continue

    # If all models failed, try to list available models
    discovered_models = list_gemini_models(api_key, timeout=15)
    available_models = [m for m in discovered_models if m not in candidate_models]
    for avail_model in available_models[:3]:  # Try first 3 available
        try:
            model_instance = genai.GenerativeModel(avail_model)
            response = model_instance.generate_content(prompt, generation_config=genai.types.GenerationConfig(
                temperature=0.0,
                max_output_tokens=1200,
                top_p=0.95,
                top_k=40,
            ))
            return parse_gemini_response_from_genai(response), avail_model
        except Exception:
            continue

    if last_error:
        raise last_error
    raise ValueError("Gemini model not found or unavailable. Tried models: " + ", ".join(candidate_models + available_models[:3]))
    raise RuntimeError(error_message)


def fetch_page_html(url: str, timeout: int = DEFAULT_TIMEOUT, use_playwright: bool = True) -> Tuple[str, str]:
    user_agent = random.choice(USER_AGENTS)

    if use_playwright and PLAYWRIGHT_AVAILABLE:
        try:
            with sync_playwright() as playwright:
                browser = playwright.chromium.launch(headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"])
                context = browser.new_context(
                    user_agent=user_agent,
                    viewport={"width": 1440, "height": 900},
                    locale="en-US",
                    java_script_enabled=True,
                    extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
                )
                page = context.new_page()
                page.set_default_navigation_timeout(timeout * 1000)
                page.set_default_timeout(timeout * 1000)
                page.goto(url, wait_until="networkidle")
                page.wait_for_timeout(random.randint(800, 1800))
                page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
                page.wait_for_timeout(random.randint(500, 1200))
                page.wait_for_load_state("networkidle")
                html = page.content()
                context.close()
                browser.close()
                return html, user_agent + " (Playwright)"
        except (NotImplementedError, OSError, PlaywrightError) as exc:
            # Fall back to requests if Playwright cannot run in this environment.
            response = requests.get(url, headers={"User-Agent": user_agent}, timeout=timeout)
            response.raise_for_status()
            return response.text, user_agent + f" (requests fallback due to {type(exc).__name__})"

    if use_playwright and not PLAYWRIGHT_AVAILABLE:
        raise RuntimeError("Playwright is not installed. Install it with: pip install playwright && playwright install chromium")

    response = requests.get(url, headers={"User-Agent": user_agent}, timeout=timeout)
    response.raise_for_status()
    return response.text, user_agent + " (requests fallback)"


def extract_fields_from_content(content: str) -> Dict[str, str]:
    content_text = clean_text(str(content))
    if not content_text:
        return {}

    # Normalize separators and add spaces around numbers/Arabic words
    content_text = re.sub(r'([0-9])([^0-9\s\u0600-\u06FF\w])', r'\1 \2', content_text)
    content_text = re.sub(r'([^0-9\s\u0600-\u06FF\w])([0-9])', r'\1 \2', content_text)
    content_text = re.sub(r'\s+', ' ', content_text).strip()

    fields: Dict[str, str] = {}

    # Extract birth date or date-like values
    date_match = re.search(r'\b\d{1,2}[\/\.\-]\d{1,2}[\/\.\-]\d{2,4}\b', content_text)
    if date_match:
        fields['date'] = date_match.group(0)
        content_text = content_text.replace(fields['date'], ' ')

    # Extract age if explicit
    age_match = re.search(r'(?:عمر|سنة|سنوات)\s*[:\-]?\s*(\d{1,3})\b', content_text)
    if age_match:
        fields['age'] = age_match.group(1)
        content_text = content_text.replace(age_match.group(0), ' ')
    elif re.match(r'^\d{1,3}\b', content_text):
        leading_num = re.match(r'^(\d{1,3})\b', content_text)
        if leading_num:
            fields['age'] = leading_num.group(1)
            content_text = content_text[leading_num.end():].strip()

    # Extract discount percent or price patterns
    discount_match = re.search(r'(\d{1,3})\s*%\b', content_text)
    if discount_match:
        fields['discount'] = discount_match.group(1) + '%'
        content_text = content_text.replace(discount_match.group(0), ' ')

    price_match = re.search(r'(?:\d+[\.,]?\d*\s*(?:دولار|ريال|جنيه|EGP|SAR|USD|USD|د\.إ|AED|€|\$|£))', content_text, flags=re.I)
    if price_match:
        fields['price'] = price_match.group(0).strip()
        content_text = content_text.replace(price_match.group(0), ' ')

    # Remove known noise phrases
    noise_patterns = [
        r'تاريخ الميلاد', r'ميلاد', r'اخبار اللاعب', r'اللاعب', r'خبر اللاعب',
        r'مواليد', r'تاريخ', r'العمر', r'سنة', r'سنوات', r'كمية', r'السعر', r'الخصم'
    ]
    for pattern in noise_patterns:
        content_text = re.sub(pattern, ' ', content_text, flags=re.I)

    content_text = re.sub(r'\s+', ' ', content_text).strip()

    # Infer name from remaining text
    if content_text:
        candidate = content_text
        candidate = re.sub(r'\d+', ' ', candidate).strip()
        if candidate:
            fields['name'] = candidate

    return fields


def build_export_payload(df: pd.DataFrame) -> List[Dict[str, Any]]:
    return df.fillna("").to_dict(orient="records")


def organize_data_professionally(records: List[Dict[str, Any]]) -> pd.DataFrame:
    """تنظيم البيانات بشكل احترافي مع فصل الأعمدة المناسبة"""
    if not records:
        return pd.DataFrame()

    enriched_records = []
    for record in records:
        record = dict(record)
        if 'content' in record and record.get('content'):
            parsed = extract_fields_from_content(record['content'])
            record.update(parsed)
        enriched_records.append(record)

    # قاموس لتصنيف الحقول المختلفة
    field_categories = {
        'name': ['name', 'title', 'product', 'item', 'اسم', 'عنوان', 'منتج'],
        'price': ['price', 'cost', 'amount', 'سعر', 'تكلفة', 'مبلغ'],
        'discount': ['discount', 'offer', 'sale', 'خصم', 'عرض', 'تخفيض'],
        'quantity': ['quantity', 'qty', 'amount', 'stock', 'كمية', 'عدد', 'مخزون'],
        'age': ['age', 'year', 'عمر', 'سنة'],
        'category': ['category', 'type', 'class', 'فئة', 'نوع', 'تصنيف'],
        'description': ['description', 'desc', 'details', 'وصف', 'تفاصيل'],
        'brand': ['brand', 'manufacturer', 'علامة', 'ماركة'],
        'rating': ['rating', 'rate', 'review', 'تقييم', 'تصنيف'],
        'location': ['location', 'address', 'city', 'موقع', 'عنوان', 'مدينة'],
        'phone': ['phone', 'mobile', 'contact', 'هاتف', 'جوال', 'اتصال'],
        'email': ['email', 'mail', 'بريد', 'إيميل'],
        'website': ['website', 'url', 'link', 'موقع', 'رابط'],
        'date': ['date', 'time', 'تاريخ', 'وقت'],
        'status': ['status', 'state', 'حالة', 'وضع']
    }

    # إنشاء DataFrame منظم
    organized_data = []
    for record in enriched_records:
        organized_record = {}
        for category, keywords in field_categories.items():
            for key, value in record.items():
                key_lower = key.lower()
                if any(keyword.lower() in key_lower for keyword in keywords):
                    organized_record[category] = str(value).strip()
                    break

        additional_info = {}
        for key, value in record.items():
            key_lower = key.lower()
            is_categorized = False
            for category, keywords in field_categories.items():
                if any(keyword.lower() in key_lower for keyword in keywords):
                    is_categorized = True
                    break
            if not is_categorized and key_lower not in ['index', 'type', 'content']:
                additional_info[key] = str(value).strip()

        if additional_info:
            organized_record['معلومات_إضافية'] = json.dumps(additional_info, ensure_ascii=False)

        organized_data.append(organized_record)

    df = pd.DataFrame(organized_data)

    preferred_order = [
        'name', 'category', 'brand', 'price', 'discount', 'quantity',
        'age', 'rating', 'description', 'location', 'phone', 'email',
        'website', 'date', 'status', 'معلومات_إضافية'
    ]
    existing_columns = [col for col in preferred_order if col in df.columns]
    other_columns = [col for col in df.columns if col not in preferred_order]
    df = df[existing_columns + other_columns]

    column_translations = {
        'name': 'الاسم',
        'price': 'السعر',
        'discount': 'الخصم',
        'quantity': 'الكمية',
        'age': 'العمر',
        'category': 'الفئة',
        'description': 'الوصف',
        'brand': 'العلامة التجارية',
        'rating': 'التقييم',
        'location': 'الموقع',
        'phone': 'الهاتف',
        'email': 'البريد الإلكتروني',
        'website': 'الموقع الإلكتروني',
        'date': 'التاريخ',
        'status': 'الحالة',
        'معلومات_إضافية': 'معلومات إضافية'
    }
    df.columns = [column_translations.get(col, col) for col in df.columns]

    return df.fillna('')


def extract_tables_from_html(html: str) -> List[Dict[str, Any]]:
    """Extract data from HTML tables as a list of dictionaries."""
    soup = BeautifulSoup(html, 'html.parser')
    tables = soup.find_all('table')
    all_records = []

    for table in tables:
        headers = []
        rows = table.find_all('tr')

        if not rows:
            continue

        # Extract headers
        header_row = rows[0]
        headers = [th.get_text(strip=True) for th in header_row.find_all(['th', 'td'])]

        # If no headers, create generic ones
        if not headers or all(not h for h in headers):
            headers = [f"column_{i+1}" for i in range(len(header_row.find_all(['th', 'td'])))]

        # Extract data rows
        for row in rows[1:]:
            cells = row.find_all(['td', 'th'])
            if len(cells) == len(headers):
                record = {headers[i]: cells[i].get_text(strip=True) for i in range(len(headers))}
                if any(record.values()):  # Only add if not empty
                    all_records.append(record)

    return all_records


def extract_structured_data(html: str) -> List[Dict[str, Any]]:
    """Extract JSON-LD and microdata from HTML."""
    soup = BeautifulSoup(html, 'html.parser')
    records = []

    # JSON-LD
    json_scripts = soup.find_all('script', type='application/ld+json')
    for script in json_scripts:
        try:
            data = json.loads(script.string)
            if isinstance(data, list):
                records.extend(data)
            elif isinstance(data, dict):
                records.append(data)
        except (json.JSONDecodeError, TypeError):
            pass

    # Microdata
    items = soup.find_all(attrs={"itemtype": True})
    for item in items:
        record = {}
        for prop in item.find_all(attrs={"itemprop": True}):
            record[prop.get('itemprop')] = prop.get_text(strip=True)
        if record:
            records.append(record)

    return records


def extract_lists_from_html(html: str) -> List[Dict[str, Any]]:
    """Extract data from HTML lists (ul, ol) as a list of dictionaries."""
    soup = BeautifulSoup(html, 'html.parser')
    records = []

    # Find all lists
    lists = soup.find_all(['ul', 'ol'])
    for list_idx, list_elem in enumerate(lists):
        items = list_elem.find_all('li')
        for item_idx, item in enumerate(items):
            text = item.get_text(strip=True)
            if text and len(text) > 10:  # Filter meaningful content
                records.append({
                    "content": text,
                    "list_type": list_elem.name,
                    "list_index": list_idx + 1,
                    "item_index": item_idx + 1
                })

    return records


def extract_divs_from_html(html: str) -> List[Dict[str, Any]]:
    """Extract data from divs with common data classes."""
    soup = BeautifulSoup(html, 'html.parser')
    records = []

    # Common data containers
    data_selectors = [
        'div.product', 'div.item', 'div.card', 'div.article',
        'div.post', 'div.entry', 'div.listing', 'div.record',
        '.product', '.item', '.card', '.article', '.post'
    ]

    for selector in data_selectors:
        elements = soup.select(selector)
        for i, elem in enumerate(elements):
            text = elem.get_text(strip=True)
            if text and len(text) > 20:  # Filter meaningful content
                records.append({
                    "content": text,
                    "type": selector.replace('.', '').replace('div', ''),
                    "index": i + 1
                })

    return records
    records = []

    # Common data containers
    data_selectors = [
        'div.product', 'div.item', 'div.card', 'div.article',
        'div.post', 'div.entry', 'div.listing', 'div.record',
        '.product', '.item', '.card', '.article', '.post'
    ]

    for selector in data_selectors:
        elements = soup.select(selector)
        for i, elem in enumerate(elements):
            text = elem.get_text(strip=True)
            if text and len(text) > 20:  # Filter meaningful content
                records.append({
                    "content": text,
                    "type": selector.replace('.', '').replace('div', ''),
                    "index": i + 1
                })

    return records


def extract_meta_data(html: str) -> Dict[str, str]:
    """Extract meta information from HTML."""
    soup = BeautifulSoup(html, 'html.parser')
    meta = {}

    # Title
    title = soup.find('title')
    if title:
        meta['title'] = title.get_text(strip=True)

    # Meta tags
    for tag in soup.find_all('meta'):
        name = tag.get('name') or tag.get('property')
        content = tag.get('content')
        if name and content:
            meta[name] = content

    return meta


def hybrid_extract_data(html: str, gemini_records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Combine multiple extraction methods for comprehensive data."""
    html_records = extract_tables_from_html(html)
    list_records = extract_lists_from_html(html)
    div_records = extract_divs_from_html(html)
    structured_records = extract_structured_data(html)

    # Combine all sources
    all_records = html_records + list_records + div_records + structured_records + gemini_records

    # Remove duplicates based on content similarity
    unique_records = []
    seen = set()
    for record in all_records:
        # Create a hash of sorted key-value pairs
        record_str = str(sorted(record.items()))
        if record_str not in seen and record:
            seen.add(record_str)
            unique_records.append(record)

    return unique_records


def generate_excel_from_dataframe(df: pd.DataFrame, sheet_name: str = "data") -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("Install openpyxl to enable Excel export.")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=str(header))
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4F94CD")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 4, 50)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


# ---------------------------------------------------------------------------
# UI helpers
# ---------------------------------------------------------------------------

def render_page_style() -> None:
    st.set_page_config(page_title="Universal Data Extraction Portal", page_icon="🔍", layout="wide")
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Cairo:wght@300;400;500;600;700&display=swap');
        html, body, [class*="css"] {
            font-family: 'Cairo', sans-serif;
        }
        .stApp {
            background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%);
            color: #f8fafc;
        }
        .hero-card, .result-card {
            background: rgba(15, 23, 42, 0.92);
            border: 1px solid rgba(148, 163, 184, 0.15);
            border-radius: 24px;
            padding: 28px;
            box-shadow: 0 30px 90px rgba(15, 23, 42, 0.35);
        }
        .hero-card h1 {
            color: #a5b4fc;
        }
        .metric-card {
            background: linear-gradient(135deg, #4f46e5 0%, #22d3ee 100%);
            color: white;
            border-radius: 18px;
            padding: 18px;
            text-align: center;
            margin-bottom: 12px;
        }
        .stButton > button {
            background: #1d4ed8;
            color: white;
            border-radius: 100px;
            padding: 0.9rem 1.6rem;
            font-weight: 700;
            transition: transform 0.2s ease;
        }
        .stButton > button:hover {
            transform: translateY(-2px);
        }
        .stDownloadButton > button {
            border-radius: 100px;
        }
        .stTextArea textarea {
            font-family: 'Courier New', monospace;
        }
        .stDataFrame, .stDataEditor {
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.12);
            border: 1px solid rgba(148, 163, 184, 0.2);
        }
        .stDataFrame table, .stDataEditor table {
            border-collapse: collapse;
            width: 100%;
            font-size: 14px;
        }
        .stDataFrame th, .stDataEditor th {
            background: linear-gradient(135deg, #4f46e5 0%, #22d3ee 100%);
            color: white;
            font-weight: 600;
            padding: 16px 12px;
            text-align: center;
            border-bottom: 2px solid rgba(255, 255, 255, 0.2);
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        .stDataFrame td, .stDataEditor td {
            padding: 12px;
            border-bottom: 1px solid rgba(148, 163, 184, 0.15);
            text-align: center;
            transition: background-color 0.2s ease;
        }
        .stDataFrame tr:nth-child(even), .stDataEditor tr:nth-child(even) {
            background-color: rgba(15, 23, 42, 0.03);
        }
        .stDataFrame tr:hover, .stDataEditor tr:hover {
            background-color: rgba(79, 70, 229, 0.08);
            transform: scale(1.01);
            transition: all 0.2s ease;
        }
        .stDataFrame tr:hover td, .stDataEditor tr:hover td {
            border-bottom-color: rgba(79, 70, 229, 0.3);
        }
        /* تخصيص ألوان للأعمدة المختلفة */
        .stDataFrame td:nth-child(1), .stDataEditor td:nth-child(1) { /* الاسم */
            background: linear-gradient(90deg, rgba(34, 211, 238, 0.08) 0%, rgba(34, 211, 238, 0.02) 100%);
            border-left: 3px solid #22d3ee;
        }
        .stDataFrame td:nth-child(2), .stDataEditor td:nth-child(2) { /* الفئة */
            background: linear-gradient(90deg, rgba(79, 70, 229, 0.08) 0%, rgba(79, 70, 229, 0.02) 100%);
            border-left: 3px solid #4f46e5;
        }
        .stDataFrame td:nth-child(3), .stDataEditor td:nth-child(3) { /* السعر */
            background: linear-gradient(90deg, rgba(34, 197, 94, 0.08) 0%, rgba(34, 197, 94, 0.02) 100%);
            border-left: 3px solid #22c55e;
            font-weight: 700;
            color: #15803d;
        }
        .stDataFrame td:nth-child(4), .stDataEditor td:nth-child(4) { /* الخصم */
            background: linear-gradient(90deg, rgba(239, 68, 68, 0.08) 0%, rgba(239, 68, 68, 0.02) 100%);
            border-left: 3px solid #ef4444;
            font-weight: 700;
            color: #dc2626;
        }
        .stDataFrame td:nth-child(5), .stDataEditor td:nth-child(5) { /* الكمية */
            background: linear-gradient(90deg, rgba(245, 158, 11, 0.08) 0%, rgba(245, 158, 11, 0.02) 100%);
            border-left: 3px solid #f59e0b;
            font-weight: 600;
        }
        /* تحسين عرض النصوص الطويلة مع tooltip */
        .stDataFrame td, .stDataEditor td {
            max-width: 180px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
            position: relative;
        }
        .stDataFrame td:hover::after, .stDataEditor td:hover::after {
            content: attr(title);
            position: absolute;
            bottom: 100%;
            left: 50%;
            transform: translateX(-50%);
            background: rgba(0, 0, 0, 0.9);
            color: white;
            padding: 8px 12px;
            border-radius: 6px;
            font-size: 12px;
            white-space: normal;
            word-wrap: break-word;
            max-width: 300px;
            z-index: 1000;
            pointer-events: none;
            opacity: 1;
            transition: opacity 0.3s ease;
        }
        /* تحسين الخطوط والتباعد */
        .stDataFrame, .stDataEditor {
            font-family: 'Cairo', 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.5;
        }
        /* إضافة أيقونات للحقول المهمة */
        .stDataFrame th::before, .stDataEditor th::before {
            content: '';
            display: inline-block;
            width: 16px;
            height: 16px;
            margin-right: 8px;
            vertical-align: middle;
        }
        .stDataFrame th:nth-child(1)::before, .stDataEditor th:nth-child(1)::before {
            background: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' fill='white' viewBox='0 0 24 24'%3E%3Cpath d='M12 2C13.1 2 14 2.9 14 4C14 5.1 13.1 6 12 6C10.9 6 10 5.1 10 4C10 2.9 10.9 2 12 2ZM21 9V7L15 1H5C3.89 1 3 1.89 3 3V21C3 22.11 3.89 23 5 23H19C20.11 23 21 22.11 21 21V9M19 9H14V4H19V9Z'/%3E%3C/svg%3E") no-repeat center;
            background-size: contain;
        }
        .stDataFrame th:nth-child(3)::before, .stDataEditor th:nth-child(3)::before {
            background: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' fill='white' viewBox='0 0 24 24'%3E%3Cpath d='M7,15H9C9,16.08 10.37,17 12,17C13.63,17 15,16.08 15,15C15,13.9 13.96,13.5 11.76,12.97C9.64,12.44 7,11.78 7,9C7,7.21 8.47,5.69 10.5,5.18V3H13.5V5.18C15.53,5.69 17,7.21 17,9H15C15,7.92 13.63,7 12,7C10.37,7 9,7.92 9,9C9,10.1 10.04,10.5 12.24,11.03C14.36,11.56 17,12.22 17,15C17,16.79 15.53,18.31 13.5,18.82V21H10.5V18.82C8.47,18.31 7,16.79 7,15Z'/%3E%3C/svg%3E") no-repeat center;
            background-size: contain;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_sidebar() -> Tuple[str, str, bool, int]:
    st.sidebar.title("⚙️ إعدادات المحرك")
    api_key = st.sidebar.text_input("Gemini API Key", type="password", help="ضع مفتاح Google Gemini API الخاص بك هنا. تأكد من لصق المفتاح فقط، وليس رسالة خطأ أو نصاً آخر.")
    model = st.sidebar.text_input("Gemini Model", value=DEFAULT_GEMINI_MODEL, help="يمكنك وضع gemini-1.5-flash أو gemini-1.5-pro إذا كان متاحاً.")
    enable_fetch = st.sidebar.checkbox("استخدام Playwright Fetcher", value=True)
    timeout = st.sidebar.slider("Timeout (seconds)", min_value=30, max_value=90, value=DEFAULT_TIMEOUT, help="وقت الانتظار حتى يتم تحميل الصفحة بالكامل.")
    st.sidebar.markdown("---")
    st.sidebar.markdown(
        "**الاستخراج الذكي:** جداول HTML + Gemini AI + قوائم تلقائية  \n        **التصدير:** CSV, Excel, JSON, TXT  \n        **المصدر:** URL أو HTML أو ملفات بيانات مباشرة  \n        **النموذج:** Gemini API ديناميكي بدون قواعد ثابتة"
    )
    return api_key, model, enable_fetch, timeout


def display_metrics(source_name: str, page_text: str, records: List[Dict[str, Any]], df: pd.DataFrame) -> None:
    st.markdown(f"<div class='result-card'><h2>✅ نتائج الاستخراج من: {source_name}</h2></div>", unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"<div class='metric-card'><strong>الكلمات</strong><br>{len(page_text.split())}</div>", unsafe_allow_html=True)
    with col2:
        st.markdown(f"<div class='metric-card'><strong>الحروف</strong><br>{len(page_text)}</div>", unsafe_allow_html=True)
    with col3:
        st.markdown(f"<div class='metric-card'><strong>السجلات</strong><br>{len(records)}</div>", unsafe_allow_html=True)
    with col4:
        st.markdown(f"<div class='metric-card'><strong>الأعمدة</strong><br>{len(df.columns) if not df.empty else 0}</div>", unsafe_allow_html=True)


def render_export_section(source_name: str, df: pd.DataFrame, raw_text: str, json_records: List[Dict[str, Any]]) -> None:
    with st.expander("📥 تصدير البيانات"):
        if not df.empty:
            st.download_button(
                "📄 تنزيل CSV",
                data=df.to_csv(index=False, encoding="utf-8"),
                file_name=f"{source_name}_records.csv",
                mime="text/csv",
            )
            if OPENPYXL_AVAILABLE:
                st.download_button(
                    "📊 تنزيل Excel",
                    data=generate_excel_from_dataframe(df, sheet_name="records"),
                    file_name=f"{source_name}_records.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            st.download_button(
                "🔗 تنزيل JSON",
                data=json.dumps(json_records, ensure_ascii=False, indent=2),
                file_name=f"{source_name}_records.json",
                mime="application/json",
            )
        if raw_text:
            st.download_button(
                "📝 تنزيل نص خام",
                data=raw_text.encode("utf-8"),
                file_name=f"{source_name}_raw.txt",
                mime="text/plain",
            )
        if not df.empty:
            st.code(json.dumps(build_export_payload(df), ensure_ascii=False, indent=2), language="json")


def show_results(source_name: str, raw_text: str, records: List[Dict[str, Any]], df: pd.DataFrame) -> None:
    display_metrics(source_name, raw_text, records, df)
    tabs = st.tabs(["📒 السجلات المستخرجة", "📄 النص الخام", "📊 نظرة على الجدول", "📥 التصدير"])

    with tabs[0]:
        if records:
            st.markdown("### 📊 البيانات المستخرجة بشكل احترافي")
            st.markdown("✅ **تم تنظيم البيانات تلقائياً في أعمدة منفصلة لكل نوع من المعلومات**")
            st.markdown("🎯 **المميزات:** فصل الأسماء، الأسعار، الخصومات، الكميات في أعمدة مستقلة")

            # تنظيم البيانات بشكل احترافي
            organized_df = organize_data_professionally(records)

            # عرض إحصائيات محسنة
            st.markdown("---")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📋 السجلات المستخرجة", len(organized_df))
            with col2:
                st.metric("📊 الأعمدة المنظمة", len(organized_df.columns))
            with col3:
                categorized_cols = len([col for col in organized_df.columns if col != 'معلومات إضافية'])
                st.metric("🏷️ الحقول المصنفة", categorized_cols)
            with col4:
                completeness = int((organized_df.notna().sum().sum() / (len(organized_df) * len(organized_df.columns))) * 100) if len(organized_df) > 0 else 0
                st.metric("✅ اكتمال البيانات", f"{completeness}%")

            st.markdown("---")

            # جدول قابل للتحرير مع إعدادات محسنة
            st.markdown("### ✏️ تحرير البيانات المهنية")
            st.markdown("💡 **نصائح للاستخدام:**")
            st.markdown("- كل عمود يحتوي على نوع بيانات محدد (اسم، سعر، خصم، كمية، إلخ)")
            st.markdown("- يمكنك النقر على أي خلية لتحريرها مباشرة")
            st.markdown("- استخدم الزر (+) لإضافة صفوف جديدة")
            st.markdown("- استخدم الزر (×) لحذف صفوف غير مرغوبة")

            # إنشاء إعدادات الأعمدة الديناميكية
            column_config = {}
            for col in organized_df.columns:
                if col == 'الاسم':
                    column_config[col] = st.column_config.TextColumn(col, width="medium", help="اسم المنتج أو العنصر")
                elif col == 'السعر':
                    column_config[col] = st.column_config.TextColumn(col, width="small", help="السعر الأصلي")
                elif col == 'الخصم':
                    column_config[col] = st.column_config.TextColumn(col, width="small", help="قيمة الخصم أو العرض")
                elif col == 'الكمية':
                    column_config[col] = st.column_config.NumberColumn(col, width="small", help="الكمية المتاحة", min_value=0)
                elif col == 'العمر':
                    column_config[col] = st.column_config.NumberColumn(col, width="small", help="العمر أو السنة", min_value=0)
                elif col == 'التقييم':
                    column_config[col] = st.column_config.NumberColumn(col, width="small", help="التقييم أو التصنيف", min_value=0.0, max_value=5.0, step=0.1)
                elif col in ['الوصف', 'معلومات إضافية']:
                    column_config[col] = st.column_config.TextColumn(col, width="large", help="وصف مفصل أو معلومات إضافية")
                elif col in ['الهاتف', 'البريد الإلكتروني', 'الموقع الإلكتروني']:
                    column_config[col] = st.column_config.TextColumn(col, width="medium", help=f"معلومات {col}")
                else:
                    column_config[col] = st.column_config.TextColumn(col, width="medium")

            edited_df = st.data_editor(
                organized_df,
                use_container_width=True,
                num_rows="dynamic",
                key="editor_records",
                column_config=column_config
            )
            st.session_state['edited_df'] = edited_df

            # عرض ملخص سريع
            st.markdown("---")
            st.markdown("### 📈 ملخص البيانات")
            if len(organized_df) > 0:
                summary_cols = st.columns(min(len(organized_df.columns), 6))
                for i, col in enumerate(organized_df.columns[:6]):
                    with summary_cols[i % 6]:
                        non_empty = organized_df[col].notna().sum()
                        st.metric(f"{col}", f"{non_empty}/{len(organized_df)}", f"{non_empty * 100 // len(organized_df)}%")

        else:
            st.warning("⚠️ لم يتم استخراج أي كيانات منظمة. تأكد من أن الصفحة تحتوي على سجلات قابلة للتكرار.")

    with tabs[1]:
        st.text_area("النص المصدر", raw_text, height=400)

    with tabs[2]:
        if not df.empty:
            st.markdown("### 📈 الجدول الأصلي (بعرض احترافي)")
            st.markdown("✅ البيانات الخام من المصدر مع تنسيق احترافي وقابلية للتحرير")

            # إحصائيات الجدول الأصلي
            st.markdown("---")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📋 إجمالي الصفوف", len(df))
            with col2:
                st.metric("📊 عدد الأعمدة", len(df.columns))
            with col3:
                total_cells = len(df) * len(df.columns)
                filled_cells = df.notna().sum().sum()
                fill_percent = int((filled_cells / total_cells * 100)) if total_cells > 0 else 0
                st.metric("📈 نسبة الامتلاء", f"{fill_percent}%")
            with col4:
                st.metric("✅ الخلايا المملوءة", f"{filled_cells}/{total_cells}")

            st.markdown("---")

            # عرض ديناميكي قابل للتحرير للجدول الأصلي
            st.markdown("### ✏️ تحرير الجدول الأصلي")
            st.markdown("💡 **يمكنك تحرير البيانات الخام مباشرة:**")
            
            # تحديد أنواع الأعمدة تلقائياً
            original_column_config = {}
            for col in df.columns:
                # محاولة تحديد نوع البيانات
                try:
                    # محاولة تحويل إلى أرقام
                    pd.to_numeric(df[col].dropna())
                    original_column_config[col] = st.column_config.NumberColumn(col, width="small")
                except (ValueError, TypeError):
                    # إذا كانت نصوص
                    if df[col].astype(str).str.len().max() > 50:
                        original_column_config[col] = st.column_config.TextColumn(col, width="large")
                    else:
                        original_column_config[col] = st.column_config.TextColumn(col, width="medium")

            edited_original_df = st.data_editor(
                df.fillna(""),
                use_container_width=True,
                num_rows="dynamic",
                key="editor_original",
                column_config=original_column_config
            )
            st.session_state['edited_original_df'] = edited_original_df

            # عرض إحصائيات الأعمدة
            st.markdown("---")
            st.markdown("### 📊 إحصائيات الأعمدة")
            stats_cols = st.columns(min(len(df.columns), 5))
            for i, col in enumerate(df.columns[:5]):
                with stats_cols[i % 5]:
                    unique_count = df[col].nunique()
                    non_empty = df[col].notna().sum()
                    st.metric(f"🏷️ {col[:15]}", f"{unique_count} قيمة", f"{non_empty} مملوء")

        else:
            st.info("لا يوجد جدول أصلي لعرضه.")

    with tabs[3]:
        # Use edited df if available (prefer organized then original)
        export_df = st.session_state.get('edited_df')
        if export_df is None:
            export_df = st.session_state.get('edited_original_df')
        if export_df is None and records:
            export_df = organize_data_professionally(records)
        elif export_df is None:
            export_df = df

        export_records = export_df.to_dict(orient="records") if not export_df.empty else records
        render_export_section(source_name, export_df, raw_text, export_records)


def main() -> None:
    render_page_style()
    api_key, model, enable_fetch, timeout = render_sidebar()

    st.markdown(
        """
        <div class='hero-card'>
            <h1>🔍 Universal Data Extraction Portal</h1>
            <p>منصة استخراج بيانات احترافية تعتمد على Playwright، HTML parsing، و Gemini API. استخراج تلقائي من الجداول والقوائم مع ذكاء اصطناعي ديناميكي.</p>
            <p>اكتب رابط صفحة أو ارفع ملف HTML/text، ثم شاهد النتائج تتشكل تلقائياً في جدول ديناميكي قابل للتصدير.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    api_key = normalize_api_key(api_key)
    tab_url, tab_file = st.tabs(["🌐 استخراج من رابط", "📁 رفع ملف"])

    with tab_url:
        url = st.text_input("أدخل رابط الصفحة هنا", placeholder="https://example.com")
        action = st.button("🚀 ابدأ الاستخراج")

        if action:
            if not validate_url(url):
                st.error("الرابط غير صالح. تأكد من أنه يبدأ بـ http:// أو https://")
            elif not api_key:
                st.error("أدخل مفتاح Gemini API صالحاً في الشريط الجانبي.")
            elif not is_valid_api_key(api_key):
                st.error("المفتاح غير صالح. ضع مفتاح API الحقيقي فقط، مثل AIza... أو ya29... لا تنسخ رسالة خطأ كاملة.")
            else:
                if enable_fetch and not PLAYWRIGHT_AVAILABLE:
                    st.warning("Playwright غير مثبت. سيتم استخدام طريقة fallback بواسطة requests بدلاً من متصفح خفي. لتفعيل Playwright الكامل، ثبت playwright ثم شغّل playwright install chromium.")
                with st.spinner("جاري فتح الصفحة داخل متصفح خفي وفحص المحتوى..."):
                    try:
                        html_content, user_agent = fetch_page_html(url, timeout, use_playwright=enable_fetch)
                        raw_text = simplify_html_to_text(html_content)
                        if is_blocked_page(raw_text):
                            st.error("يبدو أن الصفحة محمية أو غير قابلة للوصول. حاول استخدام رابط آخر أو تحقق من إعدادات الوصول.")
                        else:
                            st.success(f"تم تحميل الصفحة بنجاح باستخدام User-Agent: {user_agent}")
                            with st.spinner("جاري إرسال المحتوى إلى Gemini لتحليل الكيانات..."):
                                gemini_records, used_model = call_gemini_api(raw_text, api_key, model, timeout)
                                if used_model != model:
                                    st.info(f"تم استخدام النموذج البديل: {used_model} لأن النموذج الأصلي غير متاح.")
                                # Combine HTML extraction with Gemini results
                                records = hybrid_extract_data(html_content, gemini_records)
                                df = pd.DataFrame(records) if records else pd.DataFrame()
                                show_results(url, raw_text=raw_text, records=records, df=df)
                    except PlaywrightTimeoutError:
                        st.error("انتهت مهلة تحميل الصفحة. جرب زيادة قيمة Timeout أو تحقق من الرابط.")
                    except PlaywrightError as exc:
                        st.error(f"خطأ في Playwright: {exc}")
                    except requests.HTTPError as exc:
                        st.error(f"استجابة Gemini فشلت: {exc}")
                    except Exception as exc:
                        st.error(f"حدث خطأ غير متوقع: {exc}")

    with tab_file:
        uploaded = st.file_uploader("اختر ملف HTML أو TXT لاختبار استخراج البيانات", type=["html", "htm", "txt"])
        if uploaded:
            if uploaded.size > MAX_FILE_SIZE:
                st.error("الملف كبير جداً. الحد الأقصى هو 20MB.")
            elif not api_key:
                st.error("أدخل مفتاح Gemini API في الشريط الجانبي.")
            else:
                if st.button("🚀 استخراج من الملف"):
                    with st.spinner("جاري قراءة الملف وإرساله إلى Gemini..."):
                        try:
                            content = uploaded.read().decode("utf-8", errors="ignore")
                            raw_text = simplify_html_to_text(content) if uploaded.name.lower().endswith((".html", ".htm")) else clean_text(content)
                            if is_blocked_page(raw_text):
                                st.error("نص الملف لا يبدو صالحاً أو يحتوي على محتوى غير قابل للاستخراج.")
                            else:
                                gemini_records, used_model = call_gemini_api(raw_text, api_key, model, timeout)
                                if used_model != model:
                                    st.info(f"تم استخدام النموذج البديل: {used_model} لأن النموذج الأصلي غير متاح.")
                                # For HTML files, also extract tables directly
                                html_records = extract_tables_from_html(content) if uploaded.name.lower().endswith((".html", ".htm")) else []
                                records = hybrid_extract_data(content, gemini_records + html_records)
                                df = pd.DataFrame(records) if records else pd.DataFrame()
                                show_results(uploaded.name, raw_text=raw_text, records=records, df=df)
                        except Exception as exc:
                            st.error(f"فشل استخراج الملف: {exc}")

    st.markdown("---")
    st.markdown("<p style='text-align: center; color: #94a3b8;'>جاهز للإنتاج: تصدير سريع، جداول ديناميكية، ومرونة Gemini الديناميكية.</p>", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
